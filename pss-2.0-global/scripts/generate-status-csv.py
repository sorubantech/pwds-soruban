"""Generate a management-friendly CSV of all PSS 2.0 screen statuses.

Reads .claude/screen-tracker/REGISTRY.md, dedupes by registry ID (preferring
COMPLETED rows when duplicates exist), and writes a CSV with five user-friendly
columns: Module / Screen Name / Development Status / Dev Testing Status / QA Testing Status,
plus context columns (ID, Type, Priority, Notes).
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import os

REGISTRY = Path(".claude/screen-tracker/REGISTRY.md")

# When SUFFIX env var is set, append it to filenames (used to dodge file locks).
_SUFFIX = os.environ.get("SUFFIX", "")
OUT_CSV = Path(f".claude/screen-tracker/SCREEN-STATUS-OVERVIEW{_SUFFIX}.csv")
OUT_HTML = Path(f".claude/screen-tracker/SCREEN-STATUS-OVERVIEW{_SUFFIX}.html")
OUT_XLSX = Path(f".claude/screen-tracker/SCREEN-STATUS-OVERVIEW{_SUFFIX}.xlsx")

TEAM_MEMBERS = ["Karthick", "Saranya", "Kavin", "Sangeetha"]

DEV_STATUS_OPTIONS = [
    "Completed",
    "Completed - Needs Fix",
    "Ready to Build",
    "In Progress (Partial)",
    "Not Started",
    "Upcoming (Dashboard)",
    "Upcoming (Config)",
    "Upcoming (Mobile)",
    "Upcoming",
]

TESTING_STATUS_OPTIONS = [
    "Not Started",
    "Pending",
    "In Testing",
    "Passed",
    "Failed",
    "Issues Found",
    "Blocked - Awaiting Fix",
    "N/A (Upcoming)",
]

PRIORITY_OPTIONS = ["High", "Medium", "Low"]

# Map raw P1-P5 codes from REGISTRY.md to UX-friendly priority labels.
PRIORITY_LABEL = {
    "P1-Setup": "High",
    "P2-Core": "High",
    "P3-Business": "Medium",
    "P4-Advanced": "Low",
    "P5-Alignment": "Medium",
}


def priority_label(raw: str) -> str:
    raw = raw.strip()
    if raw in PRIORITY_LABEL:
        return PRIORITY_LABEL[raw]
    # Unknown/empty/em-dash priorities (typically skip rows) → Low
    if raw in {"", "—", "-", "–"}:
        return "Low"
    return "Medium"


def dev_done_by(dev_status: str) -> str:
    """Pre-fill 'Development Done By' for any screen whose dev work is finished."""
    if dev_status in {"COMPLETED", "NEEDS_FIX"}:
        return "Karthick"
    return ""


# Status priority for dedup (higher = preferred when same ID appears twice)
STATUS_PRIORITY = {
    "COMPLETED": 100,
    "NEEDS_FIX": 90,
    "IN_PROGRESS": 85,
    "PARTIALLY_COMPLETED": 84,
    "PROMPT_READY": 70,
    "PARTIAL": 50,
    "NEW": 40,
    "SKIP_DASHBOARD": 10,
    "SKIP_CONFIG": 10,
    "SKIP_MOBILE": 10,
    "SKIP": 10,
}

# Friendly labels for management
DEV_STATUS_LABEL = {
    "COMPLETED": "Completed",
    "NEEDS_FIX": "Completed - Needs Fix",
    "IN_PROGRESS": "In Progress",
    "PARTIALLY_COMPLETED": "Partially Completed",
    "PROMPT_READY": "Ready to Build",
    "PARTIAL": "In Progress (Partial)",
    "NEW": "Not Started",
    "SKIP_DASHBOARD": "Upcoming (Dashboard)",
    "SKIP_CONFIG": "Upcoming (Config)",
    "SKIP_MOBILE": "Upcoming (Mobile)",
    "SKIP": "Upcoming",
}

MODULE_HEADER = re.compile(r"^## (.+?)\s*$")
SUB_HEADER = re.compile(r"^### ")
DATA_ROW = re.compile(r"^\|")

# Module sections that contain real screen rows
INCLUDE_SECTIONS = {
    "FUNDRAISING MODULE",
    "CONTACTS MODULE",
    "COMMUNICATION MODULE",
    "ORGANIZATION MODULE",
    "CASE MANAGEMENT MODULE",
    "VOLUNTEER MODULE",
    "MEMBERSHIP MODULE",
    "GRANTS MODULE",
    "FIELD COLLECTION MODULE",
    "ADMINISTRATION MODULE",
    "SETTINGS MODULE",
    "AI INTELLIGENCE MODULE",
    "REPORTS MODULE",
    "MOBILE MODULE",
    "ROOT / LAYOUT",
}


MODULE_DISPLAY = {
    "FUNDRAISING": "Fundraising",
    "CONTACTS": "Contacts",
    "COMMUNICATION": "Communication",
    "ORGANIZATION": "Organization",
    "CASE MANAGEMENT": "Case Management",
    "VOLUNTEER": "Volunteer",
    "MEMBERSHIP": "Membership",
    "GRANTS": "Grants",
    "FIELD COLLECTION": "Field Collection",
    "ADMINISTRATION": "Administration",
    "SETTINGS": "Settings",
    "AI INTELLIGENCE": "AI Intelligence",
    "REPORTS": "Reports",
    "MOBILE": "Mobile",
    "ROOT / LAYOUT": "Root / Layout",
}


def clean_module(raw: str) -> str:
    name = raw.strip()
    if name.endswith(" MODULE"):
        name = name[: -len(" MODULE")]
    return MODULE_DISPLAY.get(name, name.title())


def parse_row(line: str) -> list[str] | None:
    parts = [p.strip() for p in line.strip().strip("|").split("|")]
    if len(parts) < 9:
        return None
    if parts[0] in {"#", ""} or set(parts[0]) <= {"-", " "}:
        return None
    return parts


def dev_testing_status(dev_status: str) -> str:
    if dev_status == "COMPLETED":
        return "Pending"
    if dev_status == "NEEDS_FIX":
        return "Blocked - Awaiting Fix"
    if dev_status.startswith("SKIP"):
        return "N/A (Upcoming)"
    return "Not Started"


def qa_testing_status(dev_status: str) -> str:
    if dev_status.startswith("SKIP"):
        return "N/A (Upcoming)"
    return "Not Started"


def short_notes(notes: str, limit: int = 250) -> str:
    notes = notes.replace("\n", " ").replace("\r", " ").strip()
    if len(notes) > limit:
        notes = notes[: limit - 3].rstrip() + "..."
    return notes


def main() -> None:
    rows_by_id: dict[str, dict] = {}
    current_module: str | None = None

    with REGISTRY.open(encoding="utf-8") as f:
        for line in f:
            m = MODULE_HEADER.match(line)
            if m:
                section_name = m.group(1).strip()
                if section_name in INCLUDE_SECTIONS:
                    current_module = clean_module(section_name)
                else:
                    current_module = None
                continue
            if SUB_HEADER.match(line):
                current_module = None
                continue
            if not current_module or not DATA_ROW.match(line):
                continue
            parts = parse_row(line)
            if parts is None:
                continue
            screen_id, screen_name, _, _, screen_type, priority, status, _, notes = parts[:9]
            if not screen_id:
                continue
            existing = rows_by_id.get(screen_id)
            new_priority = STATUS_PRIORITY.get(status, 0)
            if existing and STATUS_PRIORITY.get(existing["status"], 0) >= new_priority:
                continue
            rows_by_id[screen_id] = {
                "id": screen_id,
                "module": current_module,
                "screen": screen_name,
                "type": screen_type,
                "priority": priority,
                "status": status,
                "notes": notes,
            }

    sorted_rows = sorted(
        rows_by_id.values(),
        key=lambda r: _id_sort(r["id"]),
    )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow([
            "Screen #",
            "Module",
            "Screen Name",
            "Type",
            "Priority",
            "Development Status",
            "Development Done By",
            "Dev Testing Status",
            "Dev Tested By",
            "QA Testing Status",
            "QA Tested By",
        ])
        for r in sorted_rows:
            writer.writerow([
                r["id"],
                r["module"],
                r["screen"],
                r["type"],
                priority_label(r["priority"]),
                DEV_STATUS_LABEL.get(r["status"], r["status"]),
                dev_done_by(r["status"]),
                dev_testing_status(r["status"]),
                "",  # Dev Tested By — pick from dropdown after testing
                qa_testing_status(r["status"]),
                "",  # QA Tested By — pick from dropdown after testing
            ])

    # Roll-up summary
    by_status: dict[str, int] = {}
    for r in sorted_rows:
        label = DEV_STATUS_LABEL.get(r["status"], r["status"])
        by_status[label] = by_status.get(label, 0) + 1

    write_html(sorted_rows, by_status)
    write_xlsx(sorted_rows)

    print(f"Wrote {len(sorted_rows)} rows -> {OUT_CSV}")
    print(f"Wrote HTML view -> {OUT_HTML}")
    print(f"Wrote XLSX view -> {OUT_XLSX}")
    print("\nBy Development Status:")
    for label, count in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"  {label:35s} {count:3d}")


def html_class(label: str) -> str:
    if label == "High":
        return "high"
    if label == "Medium":
        return "medium"
    if label == "Low":
        return "low"
    if label.startswith("Completed - Needs"):
        return "needs-fix"
    if label.startswith("Completed"):
        return "completed"
    if label.startswith("Ready"):
        return "ready"
    if label.startswith("In Progress"):
        return "partial"
    if label.startswith("Not Started"):
        return "not-started"
    if label.startswith("Upcoming") or label.startswith("N/A (Upcoming"):
        return "upcoming"
    if label.startswith("Pending"):
        return "pending"
    if label.startswith("In Testing"):
        return "in-testing"
    if label.startswith("Passed"):
        return "completed"
    if label.startswith("Failed"):
        return "blocked"
    if label.startswith("Issues"):
        return "needs-fix"
    if label.startswith("Blocked"):
        return "blocked"
    return ""


# Status → (fill ARGB hex, font ARGB hex). ARGB has 'FF' alpha prefix.
XLSX_FILL = {
    "completed":   ("FFDCFCE7", "FF166534"),  # green
    "needs-fix":   ("FFFEF3C7", "FF92400E"),  # amber
    "ready":       ("FFDBEAFE", "FF1E40AF"),  # blue
    "partial":     ("FFFEF9C3", "FF854D0E"),  # yellow
    "not-started": ("FFF3F4F6", "FF4B5563"),  # gray
    "upcoming":    ("FFE0E7FF", "FF3730A3"),  # indigo (distinct from gray)
    "pending":     ("FFEDE9FE", "FF6B21A8"),  # purple
    "in-testing":  ("FFCFFAFE", "FF155E75"),  # cyan (distinct from pending purple)
    "blocked":     ("FFFEE2E2", "FF991B1B"),  # red
    # Priority palette
    "high":        ("FFFEE2E2", "FF991B1B"),  # red — needs attention
    "medium":      ("FFFEF3C7", "FF92400E"),  # amber — moderate
    "low":         ("FFE0F2FE", "FF075985"),  # sky blue — low concern, calm
}

# Each picklist value -> the color klass it should paint when chosen.
STATUS_COLOR_BY_LABEL = {
    "Completed": "completed",
    "Completed - Needs Fix": "needs-fix",
    "Ready to Build": "ready",
    "In Progress (Partial)": "partial",
    "Not Started": "not-started",
    "Upcoming (Dashboard)": "upcoming",
    "Upcoming (Config)": "upcoming",
    "Upcoming (Mobile)": "upcoming",
    "Upcoming": "upcoming",
    "Pending": "pending",
    "In Testing": "in-testing",
    "Passed": "completed",
    "Failed": "blocked",
    "Issues Found": "needs-fix",
    "Blocked - Awaiting Fix": "blocked",
    "N/A (Upcoming)": "upcoming",
    # Priority labels reuse the same color palette
    "High": "high",
    "Medium": "medium",
    "Low": "low",
}


def write_html(rows: list[dict], by_status: dict[str, int]) -> None:
    from html import escape

    today = __import__("datetime").date.today().isoformat()
    total = len(rows)

    summary_cards = "".join(
        f"<div class='card {html_class(label)}'><span class='count'>{count}</span><span class='label'>{escape(label)}</span></div>"
        for label, count in sorted(by_status.items(), key=lambda x: -x[1])
    )

    table_rows = []
    for r in rows:
        dev_label = DEV_STATUS_LABEL.get(r["status"], r["status"])
        dev_test = dev_testing_status(r["status"])
        qa_test = qa_testing_status(r["status"])
        dev_by = dev_done_by(r["status"])
        dev_by_cell = f"<span class='person'>{escape(dev_by)}</span>" if dev_by else "—"
        prio_label = priority_label(r["priority"])
        table_rows.append(
            "<tr>"
            f"<td class='id'>#{escape(r['id'])}</td>"
            f"<td>{escape(r['module'])}</td>"
            f"<td class='screen'>{escape(r['screen'])}</td>"
            f"<td>{escape(r['type'])}</td>"
            f"<td><span class='badge {html_class(prio_label)}'>{escape(prio_label)}</span></td>"
            f"<td><span class='badge {html_class(dev_label)}'>{escape(dev_label)}</span></td>"
            f"<td class='by'>{dev_by_cell}</td>"
            f"<td><span class='badge {html_class(dev_test)}'>{escape(dev_test)}</span></td>"
            f"<td class='by'>—</td>"
            f"<td><span class='badge {html_class(qa_test)}'>{escape(qa_test)}</span></td>"
            f"<td class='by'>—</td>"
            "</tr>"
        )

    html = f"""<!DOCTYPE html>
<html lang='en'>
<head>
<meta charset='UTF-8'>
<title>PSS 2.0 — Screen Status Overview</title>
<style>
:root {{
  --c-completed:  #16a34a;
  --c-needs-fix:  #d97706;
  --c-ready:      #2563eb;
  --c-partial:    #ca8a04;
  --c-not-started:#6b7280;
  --c-upcoming:   #6366f1;
  --c-pending:    #7c3aed;
  --c-blocked:    #dc2626;
  --bg:           #f9fafb;
  --fg:           #111827;
  --muted:        #6b7280;
  --border:       #e5e7eb;
  --card-bg:      #ffffff;
}}
* {{ box-sizing: border-box; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  background: var(--bg); color: var(--fg);
  margin: 0; padding: 24px;
}}
h1 {{ margin: 0 0 4px; font-size: 22px; }}
.meta {{ color: var(--muted); font-size: 13px; margin-bottom: 24px; }}
.summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 28px; }}
.card {{
  background: var(--card-bg); border: 1px solid var(--border); border-radius: 8px;
  padding: 14px 16px; display: flex; flex-direction: column; gap: 4px;
  border-left: 4px solid var(--muted);
}}
.card.completed   {{ border-left-color: var(--c-completed); }}
.card.needs-fix   {{ border-left-color: var(--c-needs-fix); }}
.card.ready       {{ border-left-color: var(--c-ready); }}
.card.partial     {{ border-left-color: var(--c-partial); }}
.card.not-started {{ border-left-color: var(--c-not-started); }}
.card.upcoming    {{ border-left-color: var(--c-upcoming); }}
.count {{ font-size: 28px; font-weight: 700; }}
.label {{ font-size: 12px; color: var(--muted); }}
table {{
  width: 100%; border-collapse: collapse; background: var(--card-bg);
  font-size: 13px; border: 1px solid var(--border); border-radius: 8px; overflow: hidden;
}}
th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border); vertical-align: top; }}
th {{ background: #f3f4f6; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em; color: var(--muted); }}
tr.module-divider td {{
  background: #eef2ff; color: #1e3a8a; font-weight: 700; font-size: 13px;
  letter-spacing: 0.04em; text-transform: uppercase; padding: 8px 12px;
}}
td.id {{ color: var(--muted); white-space: nowrap; font-variant-numeric: tabular-nums; }}
td.screen {{ font-weight: 500; }}
td.by {{ color: var(--muted); font-size: 12px; text-align: center; }}
.person {{ color: var(--fg); font-weight: 500; }}
.badge {{
  display: inline-block; padding: 3px 10px; border-radius: 999px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
  background: #e5e7eb; color: #374151;
}}
.badge.completed   {{ background: #dcfce7; color: #166534; }}
.badge.needs-fix   {{ background: #fef3c7; color: #92400e; }}
.badge.ready       {{ background: #dbeafe; color: #1e40af; }}
.badge.partial     {{ background: #fef9c3; color: #854d0e; }}
.badge.not-started {{ background: #f3f4f6; color: #4b5563; }}
.badge.upcoming    {{ background: #e0e7ff; color: #3730a3; }}
.badge.pending     {{ background: #ede9fe; color: #6b21a8; }}
.badge.in-testing  {{ background: #cffafe; color: #155e75; }}
.badge.blocked     {{ background: #fee2e2; color: #991b1b; }}
.badge.high        {{ background: #fee2e2; color: #991b1b; }}
.badge.medium      {{ background: #fef3c7; color: #92400e; }}
.badge.low         {{ background: #e0f2fe; color: #075985; }}
.legend {{ font-size: 12px; color: var(--muted); margin-top: 16px; line-height: 1.7; }}
.legend code {{ background: #f3f4f6; padding: 2px 6px; border-radius: 4px; }}
</style>
</head>
<body>
<h1>PSS 2.0 — Screen Status Overview</h1>
<p class='meta'>Total screens: {total} &nbsp;·&nbsp; Generated: {today} &nbsp;·&nbsp; Source: <code>.claude/screen-tracker/REGISTRY.md</code></p>

<div class='summary'>{summary_cards}</div>

<table>
<thead>
<tr>
  <th>#</th><th>Module</th><th>Screen</th><th>Type</th><th>Priority</th>
  <th>Development Status</th><th>Dev By</th>
  <th>Dev Testing</th><th>Dev Tested By</th>
  <th>QA Testing</th><th>QA Tested By</th>
</tr>
</thead>
<tbody>
{''.join(table_rows)}
</tbody>
</table>

<p class='legend'>
<strong>Development Status:</strong>
<span class='badge completed'>Completed</span> built &amp; verified ·
<span class='badge needs-fix'>Completed - Needs Fix</span> built but open ISSUEs ·
<span class='badge ready'>Ready to Build</span> prompt ready, awaiting build ·
<span class='badge partial'>In Progress (Partial)</span> partial code, needs alignment ·
<span class='badge not-started'>Not Started</span> nothing exists yet ·
<span class='badge upcoming'>Upcoming</span> deferred to a later phase (dashboards / mobile / config).<br>
<strong>Dev Testing:</strong>
<span class='badge pending'>Pending</span> ready for developer self-test ·
<span class='badge not-started'>Not Started</span> blocked on Dev completion ·
<span class='badge blocked'>Blocked - Awaiting Fix</span> dev raised, fix pending.<br>
<strong>QA Testing:</strong>
<span class='badge not-started'>Not Started</span> awaiting Dev Testing pass first.
</p>
</body>
</html>
"""

    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")


def write_xlsx(rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()
    ws = wb.active
    ws.title = "Screen Status"

    headers = [
        "Screen #",
        "Module",
        "Screen Name",
        "Type",
        "Priority",
        "Development Status",
        "Development Done By",
        "Dev Testing Status",
        "Dev Tested By",
        "QA Testing Status",
        "QA Tested By",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FF1F2937")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="FFE5E7EB"),
        right=Side(style="thin", color="FFE5E7EB"),
        top=Side(style="thin", color="FFE5E7EB"),
        bottom=Side(style="thin", color="FFE5E7EB"),
    )
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in rows:
        dev_label = DEV_STATUS_LABEL.get(r["status"], r["status"])
        dev_test = dev_testing_status(r["status"])
        qa_test = qa_testing_status(r["status"])
        ws.append([
            int(r["id"]) if r["id"].isdigit() else r["id"],
            r["module"],
            r["screen"],
            r["type"],
            priority_label(r["priority"]),
            dev_label,
            dev_done_by(r["status"]),
            dev_test,
            "",
            qa_test,
            "",
        ])

    # Apply colors per row
    for row_idx, r in enumerate(rows, start=2):
        dev_label = DEV_STATUS_LABEL.get(r["status"], r["status"])
        dev_test = dev_testing_status(r["status"])
        qa_test = qa_testing_status(r["status"])
        prio_label = priority_label(r["priority"])

        # Colored cells: priority (5), dev status (6), dev testing (8), qa testing (10)
        for col_idx, label in [(5, prio_label), (6, dev_label), (8, dev_test), (10, qa_test)]:
            klass = html_class(label)
            fill_argb, font_argb = XLSX_FILL.get(klass, ("FFFFFFFF", "FF111827"))
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=fill_argb)
            cell.font = Font(color=font_argb, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply border + alignment to all cells in row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx not in (5, 6, 8, 10):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Column widths tuned for readability
    widths = [10, 18, 32, 16, 14, 26, 22, 22, 22, 22, 22]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header + first column
    ws.freeze_panes = "B2"

    # Filter dropdowns on header row
    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    # ── Hidden lookup sheet for dropdown sources ──
    lookup = wb.create_sheet("Lookups")
    lookup["A1"] = "Development Status"
    lookup["B1"] = "Testing Status"
    lookup["C1"] = "Team Members"
    lookup["D1"] = "Priority"
    for i, val in enumerate(DEV_STATUS_OPTIONS, start=2):
        lookup.cell(row=i, column=1, value=val)
    for i, val in enumerate(TESTING_STATUS_OPTIONS, start=2):
        lookup.cell(row=i, column=2, value=val)
    for i, val in enumerate(TEAM_MEMBERS, start=2):
        lookup.cell(row=i, column=3, value=val)
    for i, val in enumerate(PRIORITY_OPTIONS, start=2):
        lookup.cell(row=i, column=4, value=val)
    lookup.column_dimensions["A"].width = 24
    lookup.column_dimensions["B"].width = 24
    lookup.column_dimensions["C"].width = 16
    lookup.column_dimensions["D"].width = 12
    lookup.sheet_state = "hidden"  # hide from end-users; visible via Format > Sheet > Unhide

    def _add_dv(col_letter: str, source_range: str, prompt: str, title: str, error: str) -> None:
        dv = DataValidation(
            type="list",
            formula1=f"=Lookups!{source_range}",
            allow_blank=True,
            showDropDown=False,  # openpyxl quirk: False = SHOW the dropdown arrow
        )
        dv.error = error
        dv.errorTitle = "Invalid value"
        dv.prompt = prompt
        dv.promptTitle = title
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv)

    # Priority dropdown (column E)
    _add_dv(
        "E",
        f"$D$2:$D${len(PRIORITY_OPTIONS) + 1}",
        "Pick: " + ", ".join(PRIORITY_OPTIONS),
        "Priority",
        "Choose High, Medium, or Low.",
    )

    # Status column dropdowns
    _add_dv(
        "F",
        f"$A$2:$A${len(DEV_STATUS_OPTIONS) + 1}",
        "Pick a development status from the list.",
        "Development Status",
        "Choose one of the predefined development statuses.",
    )
    _add_dv(
        "H",
        f"$B$2:$B${len(TESTING_STATUS_OPTIONS) + 1}",
        "Pick the developer testing status.",
        "Dev Testing Status",
        "Choose one of the predefined testing statuses.",
    )
    _add_dv(
        "J",
        f"$B$2:$B${len(TESTING_STATUS_OPTIONS) + 1}",
        "Pick the QA testing status.",
        "QA Testing Status",
        "Choose one of the predefined testing statuses.",
    )

    # "By" column dropdowns (team members)
    for col_letter in ("G", "I", "K"):
        _add_dv(
            col_letter,
            f"$C$2:$C${len(TEAM_MEMBERS) + 1}",
            "Pick: " + ", ".join(TEAM_MEMBERS),
            "Team member",
            "Please select a team member from the list.",
        )

    # ── Conditional formatting: cell color updates when dropdown value changes ──
    # Apply to the three status columns. Static fills set during write are
    # overridden by these rules at runtime, so picking a different value
    # immediately re-colors the cell.
    status_columns = [
        ("E", PRIORITY_OPTIONS),
        ("F", DEV_STATUS_OPTIONS),
        ("H", TESTING_STATUS_OPTIONS),
        ("J", TESTING_STATUS_OPTIONS),
    ]
    for col_letter, options in status_columns:
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        for label in options:
            klass = STATUS_COLOR_BY_LABEL.get(label, "not-started")
            fill_argb, font_argb = XLSX_FILL[klass]
            rule = CellIsRule(
                operator="equal",
                formula=[f'"{label}"'],
                fill=PatternFill("solid", fgColor=fill_argb),
                font=Font(color=font_argb, bold=True, size=10),
            )
            ws.conditional_formatting.add(cell_range, rule)

    # ── Build Summary sheet (pivot-style cross-tabs using live formulas) ──
    write_summary_sheet(wb, rows)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def write_summary_sheet(wb, rows: list[dict]) -> None:
    """Pivot-style summary on its own sheet, positioned as the first tab.

    Uses COUNTIFS formulas pointing at the 'Screen Status' sheet so totals
    update whenever a user picks a different value from a dropdown.
    """
    from datetime import date
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    last = len(rows) + 1
    sheet_ref = "'Screen Status'"

    # Insert as first tab so it opens by default
    s = wb.create_sheet("Summary", 0)
    s.sheet_view.showGridLines = False

    border = Border(
        left=Side(style="thin", color="FFE5E7EB"),
        right=Side(style="thin", color="FFE5E7EB"),
        top=Side(style="thin", color="FFE5E7EB"),
        bottom=Side(style="thin", color="FFE5E7EB"),
    )
    section_fill = PatternFill("solid", fgColor="FF1F2937")
    section_font = Font(bold=True, size=12, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FFF3F4F6")
    header_font = Font(bold=True, size=10, color="FF374151")
    total_fill = PatternFill("solid", fgColor="FFE5E7EB")
    total_font = Font(bold=True, size=10, color="FF111827")

    def section(row: int, col_span: int, text: str) -> None:
        c = s.cell(row=row, column=1, value=text)
        c.font = section_font
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
        s.row_dimensions[row].height = 26

    def header_cell(row: int, col: int, text: str) -> None:
        c = s.cell(row=row, column=col, value=text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def label_cell(row: int, col: int, text: str, bold: bool = False) -> None:
        c = s.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, size=10, color="FF111827")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border

    def value_cell(row: int, col: int, formula: str, bold: bool = False, total: bool = False) -> None:
        c = s.cell(row=row, column=col, value=formula)
        c.font = total_font if total else Font(bold=bold, size=10, color="FF111827")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        if total:
            c.fill = total_fill

    # ── Title strip ──
    s["A1"] = "PSS 2.0 — Screen Status Summary"
    s["A1"].font = Font(bold=True, size=18, color="FF111827")
    s.merge_cells("A1:K1")
    s["A2"] = f"Generated {date.today().isoformat()} · Total screens: {len(rows)} · Source: Screen Status sheet (live formulas)"
    s["A2"].font = Font(size=11, color="FF6B7280")
    s.merge_cells("A2:K2")
    s.row_dimensions[1].height = 28

    # ── KPI cards ──
    kpis = [
        ("Total", f"=COUNTA({sheet_ref}!A2:A{last})", "FF1F2937"),
        ("Completed", f'=COUNTIF({sheet_ref}!F2:F{last},"Completed")', "FF166534"),
        ("Needs Fix", f'=COUNTIF({sheet_ref}!F2:F{last},"Completed - Needs Fix")', "FF92400E"),
        ("In Progress", f'=COUNTIF({sheet_ref}!F2:F{last},"In Progress (Partial)")', "FF854D0E"),
        ("Ready to Build", f'=COUNTIF({sheet_ref}!F2:F{last},"Ready to Build")', "FF1E40AF"),
        ("Not Started", f'=COUNTIF({sheet_ref}!F2:F{last},"Not Started")', "FF4B5563"),
        ("Upcoming", f'=COUNTIF({sheet_ref}!F2:F{last},"Upcoming*")', "FF3730A3"),
        ("Dev Tested (Passed)", f'=COUNTIF({sheet_ref}!H2:H{last},"Passed")', "FF155E75"),
        ("QA Tested (Passed)", f'=COUNTIF({sheet_ref}!J2:J{last},"Passed")', "FF6B21A8"),
    ]
    kpi_label_row = 4
    kpi_value_row = 5
    for i, (label, formula, color) in enumerate(kpis):
        col = i + 1
        lc = s.cell(row=kpi_label_row, column=col, value=label)
        lc.font = Font(bold=True, size=10, color="FF6B7280")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = border
        lc.fill = header_fill
        vc = s.cell(row=kpi_value_row, column=col, value=formula)
        vc.font = Font(bold=True, size=24, color=color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = border
        s.column_dimensions[get_column_letter(col)].width = 18
    s.row_dimensions[kpi_value_row].height = 36

    # ── Section 1: Development Status by Module ──
    modules = sorted({r["module"] for r in rows})
    dev_cols = DEV_STATUS_OPTIONS  # 9 statuses

    sec_row = kpi_value_row + 2
    section(sec_row, len(dev_cols) + 2, "① Development Status by Module")
    hdr_row = sec_row + 1
    header_cell(hdr_row, 1, "Module")
    for i, status in enumerate(dev_cols):
        header_cell(hdr_row, i + 2, status)
    header_cell(hdr_row, len(dev_cols) + 2, "Total")
    s.row_dimensions[hdr_row].height = 38

    for r_idx, mod in enumerate(modules):
        row = hdr_row + 1 + r_idx
        label_cell(row, 1, mod)
        for c_idx, status in enumerate(dev_cols):
            formula = (
                f'=COUNTIFS({sheet_ref}!B2:B{last},"{mod}",'
                f'{sheet_ref}!F2:F{last},"{status}")'
            )
            value_cell(row, c_idx + 2, formula)
        # Row total
        total_formula = f'=COUNTIF({sheet_ref}!B2:B{last},"{mod}")'
        value_cell(row, len(dev_cols) + 2, total_formula, total=True)

    # Column totals row
    totals_row = hdr_row + 1 + len(modules)
    label_cell(totals_row, 1, "Total", bold=True)
    s.cell(row=totals_row, column=1).fill = total_fill
    for c_idx, status in enumerate(dev_cols):
        formula = f'=COUNTIF({sheet_ref}!F2:F{last},"{status}")'
        value_cell(totals_row, c_idx + 2, formula, total=True)
    value_cell(totals_row, len(dev_cols) + 2, f"=COUNTA({sheet_ref}!A2:A{last})", total=True)

    # ── Section 2: Testing Progress ──
    test_buckets = [
        ("Not Started",  "Not Started"),
        ("Pending",      "Pending"),
        ("In Testing",   "In Testing"),
        ("Passed",       "Passed"),
        ("Failed",       "Failed"),
        ("Issues Found", "Issues Found"),
        ("Blocked",      "Blocked - Awaiting Fix"),
        ("N/A",          "N/A (Upcoming)"),
    ]
    sec_row = totals_row + 2
    span = 1 + (len(test_buckets) * 2)
    section(sec_row, span, "② Testing Progress (Dev + QA)")
    hdr_row = sec_row + 1
    header_cell(hdr_row, 1, "Module")
    for i, (lbl, _) in enumerate(test_buckets):
        header_cell(hdr_row, 2 + i, f"Dev: {lbl}")
        header_cell(hdr_row, 2 + len(test_buckets) + i, f"QA: {lbl}")
    s.row_dimensions[hdr_row].height = 38

    for r_idx, mod in enumerate(modules):
        row = hdr_row + 1 + r_idx
        label_cell(row, 1, mod)
        for i, (_, val) in enumerate(test_buckets):
            dev_formula = (
                f'=COUNTIFS({sheet_ref}!B2:B{last},"{mod}",'
                f'{sheet_ref}!H2:H{last},"{val}")'
            )
            qa_formula = (
                f'=COUNTIFS({sheet_ref}!B2:B{last},"{mod}",'
                f'{sheet_ref}!J2:J{last},"{val}")'
            )
            value_cell(row, 2 + i, dev_formula)
            value_cell(row, 2 + len(test_buckets) + i, qa_formula)

    # Totals row
    test_totals_row = hdr_row + 1 + len(modules)
    label_cell(test_totals_row, 1, "Total", bold=True)
    s.cell(row=test_totals_row, column=1).fill = total_fill
    for i, (_, val) in enumerate(test_buckets):
        dev_total = f'=COUNTIF({sheet_ref}!H2:H{last},"{val}")'
        qa_total = f'=COUNTIF({sheet_ref}!J2:J{last},"{val}")'
        value_cell(test_totals_row, 2 + i, dev_total, total=True)
        value_cell(test_totals_row, 2 + len(test_buckets) + i, qa_total, total=True)

    # ── Section 3: Priority Breakdown ──
    sec_row = test_totals_row + 2
    section(sec_row, 5, "③ Priority Breakdown")
    hdr_row = sec_row + 1
    headers_p = ["Priority", "Completed", "In Progress (Partial)", "Other (Ready / Not Started / Upcoming)", "Total"]
    for i, h in enumerate(headers_p):
        header_cell(hdr_row, i + 1, h)
    s.row_dimensions[hdr_row].height = 38

    for r_idx, prio in enumerate(PRIORITY_OPTIONS):
        row = hdr_row + 1 + r_idx
        label_cell(row, 1, prio, bold=True)
        completed_f = (
            f'=COUNTIFS({sheet_ref}!E2:E{last},"{prio}",'
            f'{sheet_ref}!F2:F{last},"Completed")'
        )
        in_progress_f = (
            f'=COUNTIFS({sheet_ref}!E2:E{last},"{prio}",'
            f'{sheet_ref}!F2:F{last},"In Progress (Partial)")'
        )
        total_f = f'=COUNTIF({sheet_ref}!E2:E{last},"{prio}")'
        # Other = Total - Completed - In Progress
        other_f = (
            f'={total_f[1:]}-COUNTIFS({sheet_ref}!E2:E{last},"{prio}",'
            f'{sheet_ref}!F2:F{last},"Completed")-COUNTIFS({sheet_ref}!E2:E{last},"{prio}",'
            f'{sheet_ref}!F2:F{last},"In Progress (Partial)")'
        )
        value_cell(row, 2, completed_f)
        value_cell(row, 3, in_progress_f)
        value_cell(row, 4, other_f)
        value_cell(row, 5, total_f, total=True)

    # ── Section 4: Team Workload ──
    # Each column counts ONLY when (assignee = name) AND (matching status
    # column shows the corresponding outcome). Each test outcome gets its
    # own column so management can see the full picture: how much each
    # person passed vs failed vs found issues vs still pending.
    sec_row = hdr_row + 1 + len(PRIORITY_OPTIONS) + 1
    section(sec_row, 11, "④ Team Workload (counts ONLY when work is actually done)")
    hdr_row = sec_row + 1
    workload_headers = [
        "Team Member",
        "Dev Completed",
        "Dev: Passed",
        "Dev: Failed",
        "Dev: Issues",
        "Dev: Pending",
        "QA: Passed",
        "QA: Failed",
        "QA: Issues",
        "QA: Pending",
        "Total Engaged",
    ]
    for i, h in enumerate(workload_headers):
        header_cell(hdr_row, i + 1, h)
    s.row_dimensions[hdr_row].height = 42

    def countifs(by_col: str, by_val: str, status_col: str, status_vals: list[str]) -> str:
        """Build a formula that sums COUNTIFS for each status in status_vals."""
        parts = [
            f'COUNTIFS({sheet_ref}!{by_col}2:{by_col}{last},"{by_val}",'
            f'{sheet_ref}!{status_col}2:{status_col}{last},"{sv}")'
            for sv in status_vals
        ]
        return "=" + "+".join(parts)

    PENDING_STATES = ["Pending", "Not Started", "In Testing", "Blocked - Awaiting Fix"]

    for r_idx, name in enumerate(TEAM_MEMBERS):
        row = hdr_row + 1 + r_idx
        label_cell(row, 1, name, bold=True)

        # Dev Completed: G=name AND F in {Completed, Completed - Needs Fix}
        value_cell(row, 2, countifs("G", name, "F", ["Completed", "Completed - Needs Fix"]))
        # Dev: Passed / Failed / Issues / Pending — I=name AND H = each status
        value_cell(row, 3, countifs("I", name, "H", ["Passed"]))
        value_cell(row, 4, countifs("I", name, "H", ["Failed"]))
        value_cell(row, 5, countifs("I", name, "H", ["Issues Found"]))
        value_cell(row, 6, countifs("I", name, "H", PENDING_STATES))
        # QA: Passed / Failed / Issues / Pending — K=name AND J = each status
        value_cell(row, 7, countifs("K", name, "J", ["Passed"]))
        value_cell(row, 8, countifs("K", name, "J", ["Failed"]))
        value_cell(row, 9, countifs("K", name, "J", ["Issues Found"]))
        value_cell(row, 10, countifs("K", name, "J", PENDING_STATES))
        # Total = sum of cols B..J for this row
        value_cell(
            row, 11,
            f"=SUM({get_column_letter(2)}{row}:{get_column_letter(10)}{row})",
            total=True,
        )

    # Column widths
    s.column_dimensions["A"].width = 22
    for col in range(2, 22):
        s.column_dimensions[get_column_letter(col)].width = 14


def _id_sort(id_str: str) -> tuple[int, str]:
    try:
        return (int(id_str), "")
    except ValueError:
        m = re.match(r"^(\d+)", id_str)
        return (int(m.group(1)) if m else 9999, id_str)


if __name__ == "__main__":
    main()
